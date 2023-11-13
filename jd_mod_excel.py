import random
import openpyxl
import matplotlib.pyplot as grafico
from getpass import getpass

import jd_mod_utiles
import jd_mod_juego
import jd_mod_main

def estadistica_1():
    # para crear y chequear si crear o load
    
    #wb = openpyxl.Workbook() 
   # Sheet_name = wb.sheetnames
   # wb.save(filename='tarea_python_basica.xlsx')
    
    

    documento = openpyxl.load_workbook("tarea_python_basica.xlsx")
    hoja = documento['Sheet']


    hoja["A1"].value = "nombres"
    hoja["B1"].value = "ganadas"
    hoja["C1"].value = "perdidas"
    
    documento.save(filename='tarea_python_basica.xlsx')
    
    
def estadistica_siempre(jugador,ganar,perder):
    
    documento = openpyxl.load_workbook("tarea_python_basica.xlsx")
    hoja = documento['Sheet']
    
    buscador = None
    contador = 1
   # while buscador != jugador and buscador != None:
    
    
    for row in hoja.iter_rows(min_row=2, max_row=21, min_col=1, max_col=1):
        for cell in row:
            
            contador += 1
            
        # Access the cell value
        if cell.value == None:
            cell.value = jugador
            celda_ganada = str('B'+ str(contador))
            hoja[celda_ganada].value = 0
            celda_perdida = str('C'+ str(contador))
            hoja[celda_perdida].value = 0
            
            #celda =str('A'+ str(contador))
            #hoja[celda] = jugador
            if ganar == 1:
                #haz que gane
                celda_2 = str('B'+ str(contador))
                hoja[celda_2].value = 1
                break
                
                
            elif perder == 1:
                celda_3 = str('C'+ str(contador))
                hoja[celda_3].value = 1
                break
            
        elif cell.value == jugador:
            if ganar == 1:
                #haz que gane
                celda_22 = str('B'+ str(contador))
                hoja[celda_22].value += 1
                break
                
                
            elif perder == 1:
                celda_33 = str('C'+ str(contador))
                hoja[celda_33].value += 1
                break
            
            




    documento.save(filename='tarea_python_basica.xlsx')
        
        
    

    
def estadistica_borrar():
    
    documento = openpyxl.load_workbook("tarea_python_basica.xlsx")
    hoja = documento['Sheet']
    
    for row in hoja.iter_rows():
        for cell in row:
            cell.value = None  
    documento.save(filename='tarea_python_basica.xlsx')
    
def mostrar_estadistica():

    documento = openpyxl.load_workbook("tarea_python_basica.xlsx")
    Sheet_name = documento.sheetnames[0]
    hoja = documento[Sheet_name]
    lista_nombres = []
    lista_ganado = []
    lista_perdido = []
    for row in hoja.iter_rows(min_row=2, max_row=21, min_col=1, max_col=1):
        for cell in row:
            if cell.value != None:
                lista_nombres.append(cell.value)
            else:
                break
            
    
    for row in hoja.iter_rows(min_row=2, max_row=21, min_col=2, max_col=2):
        for cell in row: 
            if cell.value != None:
                lista_ganado.append(cell.value)
            else:
                break
                
                
    for row in hoja.iter_rows(min_row=2, max_row=21, min_col=3, max_col=3):
        for cell in row: 
            if cell.value != None:
                lista_perdido.append(cell.value)
            else:
                break
            
    #print(lista_nombres)
    #print(lista_score)
    #print(len(lista_nombres))
    #print(len(lista_score))
    buscado = input("introduce de qué jugador quieres las estadísticas: ")
    encontrado = None
    i = 0
    
    while encontrado != buscado:
        if len(lista_nombres) >= i+1:
            encontrado = lista_nombres[i]
            
            
        
        if encontrado == buscado:
            q_ganadas = lista_ganado[i]
            q_perdidas = lista_perdido[i]
            grafico.bar(["ganadas","perdidas"],[q_ganadas,q_perdidas])
            grafico.title("Estadística de " + lista_nombres[i] ) 
            grafico.show()
            break
            
        if len(lista_nombres) == i:
            print("jugador no encontrado")
            break
                
        else:
            i += 1
        
        if i > 33:
            print("jugador no encontrado")
            break
        

        
    


#estadistica_1()

#estadistica_borrar()
'''
estadistica_borrar()
estadistica_siempre("juan",1,0)
estadistica_siempre("Andres",1,0)
estadistica_siempre("juan",1,0)
estadistica_siempre("Andres",1,0)
estadistica_siempre("juan",1,0)
estadistica_siempre("juan",0,1)
estadistica_siempre("Andres",0,1)
estadistica_siempre("Andres",1,0)
estadistica_siempre("Andres",1,0)
estadistica_siempre("Andres",0,1)
'''
#mostrar_estadistica()